from openpyxl import load_workbook
import time

time_start=time.time()
wb = load_workbook("1000013291.xlsx")



sheet = wb.get_sheet_by_name("Sheet1")

# print(wb.get_sheet_names)
print(sheet.max_row)
print(sheet.max_column)

for rowOfCellObjects in sheet['A1':'AC148']:
    for cellObj in rowOfCellObjects:
     print(cellObj.coordinate, cellObj.value)
print(f'--- END OF ROW ---耗时：{time.time()-time_start}')